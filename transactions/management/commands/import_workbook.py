"""
Import master workbook rows into Django models (mirrors Business-Logic.md).

Expected sheets (case-insensitive): Projects, Transactions, Employees,
Old Accounts Payable (aliases supported).

Usage:
  python manage.py import_workbook /path/to/workbook.xlsx
  python manage.py import_workbook --clear /path/to/workbook.xlsx
  python manage.py import_workbook --dry-run path.xlsx

Install dependency: pip install openpyxl
"""

from __future__ import annotations

from decimal import Decimal, InvalidOperation
from datetime import date, datetime

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from accounts.models import Role, User
from employees.models import Employee, SalaryType
from projects.models import Project
from transactions.models import (
    FlowType,
    LedgerAccount,
    LedgerCategory,
    LegacyPayable,
    PartyType,
    PaymentStatus,
    Transaction,
    normalize_payment_status,
)


def _enum_coerce(raw, enum_cls, default: str | None = None) -> str | None:
    """Map imported text to a TextChoices value; fallback to default."""
    if raw is None or str(raw).strip() == "":
        return default
    s = str(raw).strip()
    low = s.lower()
    for m in enum_cls:
        if m.value.lower() == low:
            return m.value
    for m in enum_cls:
        vl = m.value.lower()
        if vl in low or low in vl:
            return m.value
    return default


def _norm(s: str | None) -> str:
    if s is None:
        return ""
    return " ".join(str(s).strip().lower().split())


def _header_map(row: tuple) -> dict[str, int]:
    m: dict[str, int] = {}
    for i, cell in enumerate(row):
        if cell is None or str(cell).strip() == "":
            continue
        key = _norm(str(cell))
        if key and key not in m:
            m[key] = i
    return m


def _pick(hmap: dict[str, int], *aliases: str) -> int | None:
    for a in aliases:
        k = _norm(a)
        if k in hmap:
            return hmap[k]
    for a in aliases:
        ak = _norm(a)
        for hk, idx in hmap.items():
            if ak in hk or hk in ak:
                return idx
    return None


def _dec(val) -> Decimal | None:
    if val is None or val == "":
        return None
    if isinstance(val, Decimal):
        return val
    if isinstance(val, (int, float)):
        return Decimal(str(val))
    s = str(val).strip().replace(",", "").replace(" ", "")
    if not s or s.lower() in ("-", "—", "n/a", "na"):
        return None
    try:
        return Decimal(s)
    except InvalidOperation:
        return None


def _to_date(val):
    if val is None or val == "":
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, (int, float)):
        try:
            from openpyxl.utils.datetime import from_excel

            return from_excel(val).date()
        except Exception:
            pass
    s = str(val).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(s[:19], fmt).date()
        except ValueError:
            continue
    return None


def _str(val, max_len: int | None = None) -> str:
    if val is None:
        return ""
    t = str(val).strip()
    if max_len is not None and len(t) > max_len:
        return t[:max_len]
    return t


def _find_sheet(wb, *names: str):
    lower_index = {s.lower(): s for s in wb.sheetnames}
    for n in names:
        if n in wb.sheetnames:
            return wb[n]
        ln = n.lower()
        if ln in lower_index:
            return wb[lower_index[ln]]
    return None


def _salary_type(raw: str) -> str:
    n = _norm(raw)
    if "daily" in n:
        return SalaryType.DAILY
    if "weekly" in n:
        return SalaryType.WEEKLY
    if "monthly" in n:
        return SalaryType.MONTHLY
    if raw and str(raw).strip():
        for st in SalaryType:
            if _norm(st.value) == n:
                return st.value
    return SalaryType.MONTHLY


class Command(BaseCommand):
    help = "Import Excel workbook (Projects, Transactions, Employees, Legacy payables) into the database."

    def add_arguments(self, parser):
        parser.add_argument(
            "workbook",
            nargs="?",
            default=None,
            help="Path to .xlsx file (e.g. prosperous future trading Est accounts system.xlsx)",
        )
        parser.add_argument(
            "--clear",
            action="store_true",
            help="Delete existing transactions, projects, employees, legacy payables before import",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Parse and report counts without writing to the database",
        )
        parser.add_argument(
            "--admin-email",
            default=None,
            help="If no users exist, create this admin user (requires --admin-password)",
        )
        parser.add_argument(
            "--admin-password",
            default=None,
            help="Password for --admin-email when bootstrapping first user",
        )

    def handle(self, *args, **options):
        try:
            import openpyxl
        except ImportError as e:
            raise CommandError(
                "openpyxl is required. Install with: pip install openpyxl"
            ) from e

        from pathlib import Path

        from django.conf import settings

        path_str = options["workbook"]
        if not path_str:
            candidates = [
                Path(settings.BASE_DIR) / "prosperous future trading Est accounts system.xlsx",
                Path(settings.BASE_DIR).parent / "prosperous future trading Est accounts system.xlsx",
                Path(settings.BASE_DIR) / "sheets" / "prosperous future trading Est accounts system.xlsx",
            ]
            path = next((p for p in candidates if p.is_file()), None)
            if not path:
                raise CommandError(
                    "No workbook path given and default file not found. Pass path to .xlsx explicitly."
                )
        else:
            path = Path(path_str).expanduser().resolve()
            if not path.is_file():
                raise CommandError(f"File not found: {path}")

        dry = options["dry_run"]
        clear = options["clear"]

        self.stdout.write(f"Loading workbook: {path}")

        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

        summary = {
            "projects": 0,
            "transactions": 0,
            "employees": 0,
            "legacy": 0,
            "warnings": 0,
        }

        if dry:
            self._dry_parse(wb, summary)
            self._print_summary(summary, dry=True)
            wb.close()
            return

        admin_email = options.get("admin_email")
        admin_password = options.get("admin_password")
        if admin_email and not admin_password:
            raise CommandError("--admin-password is required with --admin-email")
        if admin_password and not admin_email:
            raise CommandError("--admin-email is required with --admin-password")

        if clear:
            self.stdout.write(self.style.WARNING("Clearing existing data..."))
            Transaction.objects.all().delete()
            LegacyPayable.objects.all().delete()
            Project.objects.all().delete()
            Employee.objects.all().delete()

        if admin_email and not User.objects.exists():
            User.objects.create_user(
                email=admin_email,
                password=admin_password,
                role=Role.ADMIN,
                full_name="Imported Admin",
            )
            self.stdout.write(self.style.SUCCESS(f"Created admin user: {admin_email}"))

        with transaction.atomic():
            self._import_projects(wb, summary)
            self._import_transactions(wb, summary)
            self._import_employees(wb, summary)
            self._import_legacy(wb, summary)

        wb.close()
        self._print_summary(summary, dry=False)

    def _print_summary(self, summary: dict, dry: bool):
        mode = "(dry-run) " if dry else ""
        self.stdout.write(
            self.style.SUCCESS(
                f"{mode}Imported projects={summary['projects']}, "
                f"transactions={summary['transactions']}, "
                f"employees={summary['employees']}, "
                f"legacy_payables={summary['legacy']}, "
                f"warnings={summary['warnings']}"
            )
        )

    def _dry_parse(self, wb, summary: dict):
        """Count non-empty rows without DB writes."""
        ws_p = _find_sheet(wb, "Projects", "projects")
        if ws_p:
            header = next(ws_p.iter_rows(min_row=1, max_row=1, values_only=True))
            hm = _header_map(header)
            col_name = _pick(hm, "Project Name", "project name", "name")
            if col_name is not None:
                for row in ws_p.iter_rows(min_row=2, values_only=True):
                    if row and row[col_name] and str(row[col_name]).strip():
                        summary["projects"] += 1
        ws_t = _find_sheet(wb, "Transactions", "transactions")
        if ws_t:
            for i, row in enumerate(ws_t.iter_rows(min_row=2, values_only=True), start=2):
                if any(x not in (None, "") for x in (row or ())[:5]):
                    summary["transactions"] += 1
        ws_e = _find_sheet(wb, "Employees", "employees")
        if ws_e:
            header = next(ws_e.iter_rows(min_row=1, max_row=1, values_only=True))
            hm = _header_map(header)
            col_name = _pick(hm, "Employee Name", "name", "employee name")
            if col_name is not None:
                for row in ws_e.iter_rows(min_row=2, values_only=True):
                    if row and row[col_name] and str(row[col_name]).strip():
                        summary["employees"] += 1
        ws_l = _find_sheet(
            wb,
            "Old Accounts Payable",
            "old accounts payable",
            "Old_Accounts_Payable",
        )
        if ws_l:
            for row in ws_l.iter_rows(min_row=2, values_only=True):
                if row and row[0] and str(row[0]).strip():
                    summary["legacy"] += 1

    def _import_projects(self, wb, summary: dict):
        ws = _find_sheet(wb, "Projects", "projects")
        if not ws:
            self.stdout.write(self.style.WARNING("Sheet 'Projects' not found — skipping."))
            return
        header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        hm = _header_map(header)
        col_name = _pick(hm, "Project Name", "project name", "name")
        col_client = _pick(hm, "Client Name", "client name", "client")
        col_contract = _pick(
            hm,
            "Contract Value (Excl. VAT)",
            "contract value (excl. vat)",
            "contract value excl",
        )
        col_notes = _pick(hm, "Notes", "Note", "notes", "note")

        if col_name is None:
            self.stdout.write(self.style.ERROR("Projects: could not find 'Project Name' column."))
            return

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            name = _str(row[col_name], 255)
            if not name:
                continue
            client = _str(row[col_client], 255) if col_client is not None else ""
            contract = _dec(row[col_contract]) if col_contract is not None else None
            notes = ""
            if col_notes is not None:
                notes = _str(row[col_notes])
            Project.objects.update_or_create(
                name=name,
                defaults={
                    "client_name": client,
                    "contract_value_excl_vat": contract,
                    "notes": notes,
                },
            )
            summary["projects"] += 1

    def _import_transactions(self, wb, summary: dict):
        ws = _find_sheet(wb, "Transactions", "transactions")
        if not ws:
            self.stdout.write(self.style.WARNING("Sheet 'Transactions' not found — skipping."))
            return
        header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        hm = _header_map(header)

        c_date = _pick(hm, "Date", "date")
        c_desc = _pick(hm, "Description", "description")
        c_flow = _pick(hm, "Type (IN/OUT)", "type (in/out)", "type", "in/out")
        c_amt = _pick(hm, "Amount (Excl. VAT)", "amount (excl. vat)", "amount excl")
        c_acc = _pick(hm, "Account", "account")
        c_cat = _pick(hm, "Category", "category")
        c_proj = _pick(hm, "Project", "project")
        c_mat = _pick(hm, "Material Item", "material item")
        c_qty = _pick(hm, "Qty", "qty", "quantity")
        c_rate = _pick(hm, "Rate", "rate")
        c_party = _pick(hm, "Party Name", "party name")
        c_ptype = _pick(hm, "Party Type", "party type")
        c_inv = _pick(hm, "Invoice #", "invoice #", "invoice", "invoice number")
        c_taxp = _pick(hm, "Tax %", "tax %", "tax percent")
        c_pay = _pick(hm, "Payment Status", "payment status")
        c_notes = _pick(hm, "Notes", "notes", "note")

        required = [c_date, c_flow, c_amt, c_acc, c_cat]
        if any(x is None for x in required):
            self.stdout.write(
                self.style.ERROR(
                    f"Transactions: missing required columns. Found headers: {list(hm.keys())}"
                )
            )
            return

        batch: list[Transaction] = []
        batch_size = 400

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            if row[c_amt] is None or str(row[c_amt]).strip() == "":
                continue

            d = _to_date(row[c_date])
            if not d:
                summary["warnings"] += 1
                continue

            flow_raw = _str(row[c_flow])
            flow_up = flow_raw.upper()
            if flow_up not in ("IN", "OUT"):
                summary["warnings"] += 1
                continue
            flow = FlowType.IN if flow_up == "IN" else FlowType.OUT

            proj = None
            if c_proj is not None and row[c_proj] is not None and str(row[c_proj]).strip():
                pname = _str(row[c_proj], 255)
                proj, _ = Project.objects.get_or_create(
                    name=pname,
                    defaults={"client_name": "", "notes": ""},
                )

            qty = _dec(row[c_qty]) if c_qty is not None else None
            rate = _dec(row[c_rate]) if c_rate is not None else None
            tax_p = _dec(row[c_taxp]) if c_taxp is not None else None

            acc_raw = _str(row[c_acc]) if c_acc is not None else ""
            cat_raw = _str(row[c_cat]) if c_cat is not None else ""
            ptype_raw = _str(row[c_ptype]) if c_ptype is not None else ""
            pay_raw = _str(row[c_pay]) if c_pay is not None else ""
            pay_norm = normalize_payment_status(pay_raw) if pay_raw else ""
            pay_final = _enum_coerce(
                pay_norm or pay_raw, PaymentStatus, PaymentStatus.UNPAID.value
            )
            party_type_val = _enum_coerce(ptype_raw, PartyType, None) if ptype_raw else None

            t = Transaction(
                date=d,
                description=_str(row[c_desc], 500) if c_desc is not None else "",
                flow_type=flow,
                amount_excl_vat=_dec(row[c_amt]) or Decimal("0"),
                account=_enum_coerce(acc_raw, LedgerAccount, LedgerAccount.CASH.value),
                category=_enum_coerce(cat_raw, LedgerCategory, LedgerCategory.MATERIAL_PURCHASE.value),
                project=proj,
                material_item=_str(row[c_mat], 255) if c_mat is not None else "",
                qty=qty,
                rate=rate,
                party_name=_str(row[c_party], 255) if c_party is not None else "",
                party_type=party_type_val,
                invoice_number=_str(row[c_inv], 128) if c_inv is not None else "",
                tax_percent=tax_p,
                payment_status=pay_final,
                notes=_str(row[c_notes]) if c_notes is not None else "",
            )
            t.compute_tax_and_total()
            batch.append(t)
            if len(batch) >= batch_size:
                Transaction.objects.bulk_create(batch)
                summary["transactions"] += len(batch)
                batch = []

        if batch:
            Transaction.objects.bulk_create(batch)
            summary["transactions"] += len(batch)

    def _import_employees(self, wb, summary: dict):
        ws = _find_sheet(wb, "Employees", "employees")
        if not ws:
            self.stdout.write(self.style.WARNING("Sheet 'Employees' not found — skipping."))
            return
        header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        hm = _header_map(header)

        c_name = _pick(hm, "Employee Name", "employee name", "name")
        c_type = _pick(hm, "Type", "type", "salary type")
        c_monthly = _pick(hm, "Monthly Salary", "monthly salary")
        c_daily = _pick(hm, "Daily Rate", "daily rate")
        c_days = _pick(hm, "Days Worked (This Month)", "days worked", "days worked (this month)")
        c_paid = _pick(hm, "Total Salary Paid This Month", "salary paid", "total salary paid this month")
        c_adv_t = _pick(hm, "Total Advance Taken", "advance taken", "total advance taken")
        c_adv_a = _pick(hm, "Total Advance Adjusted", "advance adjusted", "total advance adjusted")
        c_prev = _pick(hm, "Previous Pending Salary", "previous pending salary")
        c_note = _pick(hm, "Note", "notes")

        if c_name is None:
            self.stdout.write(self.style.ERROR("Employees: could not find name column."))
            return

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            name = _str(row[c_name], 255)
            if not name:
                continue
            st_raw = _str(row[c_type]) if c_type is not None else SalaryType.MONTHLY
            Employee.objects.update_or_create(
                name=name,
                defaults={
                    "salary_type": _salary_type(st_raw),
                    "monthly_salary": _dec(row[c_monthly]) if c_monthly is not None else None,
                    "daily_rate": _dec(row[c_daily]) if c_daily is not None else None,
                    "days_worked": _dec(row[c_days]) if c_days is not None else None,
                    "salary_paid_this_month": _dec(row[c_paid]) if c_paid is not None else None,
                    "advance_taken": _dec(row[c_adv_t]) if c_adv_t is not None else None,
                    "advance_adjusted": _dec(row[c_adv_a]) if c_adv_a is not None else None,
                    "previous_pending_salary": _dec(row[c_prev]) if c_prev is not None else None,
                    "note": _str(row[c_note]) if c_note is not None else "",
                },
            )
            summary["employees"] += 1

    def _import_legacy(self, wb, summary: dict):
        ws = _find_sheet(
            wb,
            "Old Accounts Payable",
            "old accounts payable",
            "Old_Accounts_Payable",
        )
        if not ws:
            self.stdout.write(
                self.style.WARNING("Sheet 'Old Accounts Payable' not found — skipping.")
            )
            return
        header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        hm = _header_map(header)

        c_sup = _pick(
            hm,
            "Suppliars Names",
            "Suppliers Names",
            "supplier",
            "supplier name",
            "suppliers names",
        )
        c_payable = _pick(hm, "Total payable", "total payable")
        c_paid = _pick(hm, "Total paid", "total paid")
        c_date = _pick(hm, "Date of last paid", "date of last paid", "last paid")
        c_proj = _pick(hm, "Projects Name", "projects name", "project")
        c_stat = _pick(hm, "Status", "status note", "note")

        use_indices = c_sup is None or c_payable is None
        if use_indices:
            self.stdout.write(
                "Legacy payables: using fixed columns A–G (supplier, payable, paid, "
                "—, last paid, project, status)."
            )

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            if use_indices:
                supplier = _str(row[0], 255) if len(row) > 0 else ""
                payable = _dec(row[1]) if len(row) > 1 else None
                paid = _dec(row[2]) if len(row) > 2 else None
                dlast = _to_date(row[4]) if len(row) > 4 else None
                proj = _str(row[5], 255) if len(row) > 5 else ""
                stat = _str(row[6], 255) if len(row) > 6 else ""
            else:
                supplier = _str(row[c_sup], 255)
                if not supplier:
                    continue
                payable = _dec(row[c_payable])
                paid = _dec(row[c_paid]) if c_paid is not None else None
                dlast = _to_date(row[c_date]) if c_date is not None else None
                proj = _str(row[c_proj], 255) if c_proj is not None else ""
                if c_stat is not None:
                    stat = _str(row[c_stat], 255)
                else:
                    stat = _str(row[6], 255) if len(row) > 6 else ""

            if not supplier:
                continue
            if payable is None:
                summary["warnings"] += 1
                continue

            LegacyPayable.objects.create(
                supplier_name=supplier,
                total_payable=payable,
                total_paid=paid or Decimal("0"),
                date_last_paid=dlast,
                projects_name=proj,
                status_note=stat,
            )
            summary["legacy"] += 1
