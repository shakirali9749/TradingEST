"""
Seed the database with the first data row from each master-data sheet in trading.xlsx.

Does not modify existing application code. Report/dashboard sheets (Accounts_Summary,
Tax Report, Monthly_Report, Dashboard) are derived at runtime from transactions — no
rows are inserted for those sheets.

Usage:
  python manage.py seed_trading_xlsx_sample
  python manage.py seed_trading_xlsx_sample --workbook /path/to/trading.xlsx
  python manage.py seed_trading_xlsx_sample --clear
"""

from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.conf import settings
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction as db_transaction

from employees.models import Employee, SalaryType
from projects.models import Project
from transactions.category_utils import ensure_category_exists
from transactions.models import (
    FlowType,
    LedgerAccount,
    LegacyPayable,
    PartyType,
    PaymentStatus,
    Transaction,
    normalize_payment_status,
)


def _dec(val) -> Decimal | None:
    if val is None or val == "":
        return None
    if isinstance(val, Decimal):
        return val
    if isinstance(val, (int, float)):
        return Decimal(str(val))
    s = str(val).strip().replace(",", "")
    if not s:
        return None
    try:
        return Decimal(s)
    except InvalidOperation:
        return None


def _to_date(val) -> date | None:
    if val is None or val == "":
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    return None


def _str(val, max_len: int | None = None) -> str:
    if val is None:
        return ""
    t = str(val).strip()
    if max_len is not None:
        return t[:max_len]
    return t


def _first_data_row(ws):
    """Return row 2 values as a tuple (header is row 1)."""
    if ws.max_row < 2:
        return None
    return tuple(ws.cell(2, c).value for c in range(1, ws.max_column + 1))


def _headers(ws):
    return [_str(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]


class Command(BaseCommand):
    help = "Create sample DB objects from the first data row of trading.xlsx master sheets."

    def add_arguments(self, parser):
        parser.add_argument(
            "--workbook",
            default=None,
            help="Path to trading.xlsx (default: BASE_DIR/trading.xlsx)",
        )
        parser.add_argument(
            "--clear",
            action="store_true",
            help="Delete projects, transactions, employees, legacy payables before seeding",
        )

    def handle(self, *args, **options):
        try:
            import openpyxl
        except ImportError as e:
            raise CommandError("openpyxl is required: pip install openpyxl") from e

        path = Path(options["workbook"] or settings.BASE_DIR / "trading.xlsx").resolve()
        if not path.is_file():
            raise CommandError(f"Workbook not found: {path}")

        if options["clear"]:
            self.stdout.write(self.style.WARNING("Clearing existing master data..."))
            Transaction.objects.all().delete()
            LegacyPayable.objects.all().delete()
            Project.objects.all().delete()
            Employee.objects.all().delete()

        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        created = {}

        with db_transaction.atomic():
            created["project"] = self._seed_project(wb)
            created["transaction"] = self._seed_transaction(wb, created["project"])
            created["employee"] = self._seed_employee(wb)
            created["legacy"] = self._seed_legacy(wb)

        wb.close()
        self.stdout.write(self.style.SUCCESS(f"Seeded from {path.name}: {created}"))

    def _seed_project(self, wb) -> Project | None:
        ws = wb["Projects"] if "Projects" in wb.sheetnames else None
        if not ws:
            return None
        row = _first_data_row(ws)
        if not row:
            return None
        # A=name, B=client, C=contract excl (D–J are formulas in Excel)
        name = _str(row[0], 255)
        if not name:
            return None
        client = _str(row[1], 255) if len(row) > 1 else ""
        contract = _dec(row[2]) if len(row) > 2 else None

        project, _ = Project.objects.get_or_create(
            name=name,
            defaults={
                "reference_number": "PRJ-TMP-SEED",
                "client_name": client,
                "contract_value_excl_vat": contract,
            },
        )
        if project.reference_number.startswith("PRJ-TMP"):
            project.reference_number = f"PRJ-{project.pk:04d}"
            project.save(update_fields=["reference_number"])
        return project

    def _seed_transaction(self, wb, project: Project | None) -> Transaction | None:
        ws = wb["Transactions"] if "Transactions" in wb.sheetnames else None
        if not ws:
            return None
        row = _first_data_row(ws)
        if not row:
            return None
        # Indices per trading.xlsx headers (0-based)
        d = _to_date(row[0])
        if not d:
            return None
        description = _str(row[1], 500) if len(row) > 1 else ""
        flow_raw = _str(row[2]).upper() if len(row) > 2 else ""
        flow = FlowType.OUT if flow_raw == "OUT" else FlowType.IN
        amount = _dec(row[3]) if len(row) > 3 else None
        if amount is None:
            return None
        account = _str(row[4]) or LedgerAccount.CASH.value
        category = ensure_category_exists(_str(row[5]) or "Material Purchase")
        project_name = _str(row[6], 255) if len(row) > 6 else ""
        material_item = _str(row[7]) if len(row) > 7 else ""
        qty = _dec(row[8]) if len(row) > 8 else None
        rate = _dec(row[9]) if len(row) > 9 else None
        party_name = _str(row[10], 255) if len(row) > 10 else ""
        party_type_raw = _str(row[11]) if len(row) > 11 else ""
        invoice = _str(row[12], 128) if len(row) > 12 else ""
        tax_percent = _dec(row[13]) if len(row) > 13 else None
        # row[14]=Tax Amount, row[15]=Total Amount — formulas; computed on save
        payment_raw = _str(row[16]) if len(row) > 16 else ""
        notes_extra = _str(row[17]) if len(row) > 17 else ""

        if project is None and project_name:
            project, _ = Project.objects.get_or_create(
                name=project_name,
                defaults={
                    "reference_number": "PRJ-TMP-SEED2",
                    "client_name": "",
                },
            )
            if project.reference_number.startswith("PRJ-TMP"):
                project.reference_number = f"PRJ-{project.pk:04d}"
                project.save(update_fields=["reference_number"])

        party_type = None
        if party_type_raw:
            for pt in PartyType:
                if pt.value.lower() == party_type_raw.lower():
                    party_type = pt.value
                    break

        notes_parts = [p for p in (material_item, notes_extra) if p]
        notes = "; ".join(notes_parts)

        txn = Transaction(
            reference_number="TXN-TMP-SEED",
            date=d,
            description=description,
            flow_type=flow,
            amount_excl_vat=amount,
            account=account,
            category=category,
            project=project,
            qty=qty,
            rate=rate,
            party_name=party_name,
            party_type=party_type,
            invoice_number=invoice,
            tax_percent=tax_percent,
            payment_status=normalize_payment_status(payment_raw)
            or PaymentStatus.PAID.value,
            notes=notes,
        )
        txn.compute_tax_and_total()
        txn.save()
        canonical = f"TXN-{txn.pk:06d}"
        if txn.reference_number != canonical:
            txn.reference_number = canonical
            txn.save(update_fields=["reference_number"])
        return txn

    def _seed_employee(self, wb) -> Employee | None:
        ws = wb["Employees"] if "Employees" in wb.sheetnames else None
        if not ws:
            return None
        row = _first_data_row(ws)
        if not row:
            return None
        name = _str(row[0], 255)
        if not name:
            return None
        st_raw = _str(row[1]) if len(row) > 1 else SalaryType.MONTHLY
        st = SalaryType.MONTHLY
        if "daily" in st_raw.lower():
            st = SalaryType.DAILY
        elif "weekly" in st_raw.lower():
            st = SalaryType.WEEKLY

        # F,H,K,M are formulas — stored fields only for literals
        employee, _ = Employee.objects.update_or_create(
            name=name,
            defaults={
                "salary_type": st,
                "monthly_salary": _dec(row[2]) if len(row) > 2 else None,
                "daily_rate": _dec(row[3]) if len(row) > 3 else None,
                "days_worked": _dec(row[4]) if len(row) > 4 else None,
                "salary_paid_this_month": _dec(row[6]) if len(row) > 6 else None,
                "advance_taken": _dec(row[8]) if len(row) > 8 else None,
                "advance_adjusted": _dec(row[9]) if len(row) > 9 else None,
                "previous_pending_salary": _dec(row[11]) if len(row) > 11 else None,
                "note": _str(row[13]) if len(row) > 13 else "",
            },
        )
        return employee

    def _seed_legacy(self, wb) -> LegacyPayable | None:
        sheet = None
        for name in wb.sheetnames:
            if name.lower().replace(" ", "") in (
                "oldaccountspayable",
                "old_accounts_payable",
            ):
                sheet = name
                break
        if not sheet:
            return None
        ws = wb[sheet]
        row = _first_data_row(ws)
        if not row:
            return None
        supplier = _str(row[0], 255)
        payable = _dec(row[1]) if len(row) > 1 else None
        if not supplier or payable is None:
            return None
        paid = _dec(row[2]) if len(row) > 2 else Decimal("0")
        dlast = _to_date(row[4]) if len(row) > 4 else None
        proj = _str(row[5], 255) if len(row) > 5 else ""
        stat = _str(row[6], 255) if len(row) > 6 else ""

        return LegacyPayable.objects.create(
            supplier_name=supplier,
            total_payable=payable,
            total_paid=paid or Decimal("0"),
            date_last_paid=dlast,
            projects_name=proj,
            status_note=stat,
        )
